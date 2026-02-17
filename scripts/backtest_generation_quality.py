#!/usr/bin/env python3
"""
Backtest historical weekly plans for deterministic rule compliance.
Supports either a local XLSX export or live Google Sheets.
"""

import argparse
import json
import os
from collections import Counter
from datetime import datetime
import sys

import yaml

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from src.plan_validator import validate_plan


VIOLATION_WEIGHTS = {
    "odd_db_load": 10,
    "forbidden_split_squat": 10,
    "carry_wrong_day": 8,
    "single_arm_d_handle_saturday": 8,
    "triceps_attachment_rotation": 6,
    "biceps_grip_repeat": 6,
    "range_in_prescription": 4,
    "hold_lock_violation": 4,
}


def _normalize_sheet_name(name):
    return (name or "").strip()


def _looks_like_weekly_plan(name):
    normalized = _normalize_sheet_name(name).lower()
    return "weekly plan" in normalized


def _parse_workouts_from_xlsx_sheet(worksheet):
    workouts = []
    current = None
    day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]

    for row in worksheet.iter_rows(min_row=1, max_col=8, values_only=True):
        first = str(row[0]).strip() if row and row[0] is not None else ""
        first_lower = first.lower()
        if any(day in first_lower for day in day_names) and first_lower != "block":
            if current:
                workouts.append(current)
            current = {"date": first, "exercises": []}
            continue

        if not current:
            continue

        exercise = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not exercise or exercise.lower() == "exercise":
            continue

        current["exercises"].append(
            {
                "block": str(row[0]).strip() if len(row) > 0 and row[0] is not None else "",
                "exercise": exercise,
                "sets": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                "reps": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                "load": row[4] if len(row) > 4 else "",
                "rest": str(row[5]).strip() if len(row) > 5 and row[5] is not None else "",
                "notes": str(row[6]).strip() if len(row) > 6 and row[6] is not None else "",
            }
        )

    if current:
        workouts.append(current)
    return workouts


def _workouts_to_markdown(workouts):
    lines = []
    for workout in workouts:
        day_label = (workout.get("date") or "").upper()
        if not day_label:
            continue
        lines.append(f"## {day_label}")
        for ex in workout.get("exercises", []):
            block = (ex.get("block") or "A1").strip() or "A1"
            name = (ex.get("exercise") or "").strip()
            if not name:
                continue
            sets = (ex.get("sets") or "1").strip() or "1"
            reps = (ex.get("reps") or "1").strip() or "1"
            load = ex.get("load")
            load_str = ""
            if load not in (None, ""):
                try:
                    load_str = f"{float(load):.3f}".rstrip("0").rstrip(".")
                except (TypeError, ValueError):
                    load_str = str(load).strip()

            lines.append(f"### {block}. {name}")
            if load_str:
                lines.append(f"- {sets} x {reps} @ {load_str} kg")
            else:
                lines.append(f"- {sets} x {reps} @ 0 kg")
            lines.append(f"- **Rest:** {(ex.get('rest') or 'None').strip() if isinstance(ex.get('rest'), str) else ex.get('rest') or 'None'}")
            lines.append(f"- **Notes:** {ex.get('notes') or ''}")
        lines.append("")

    while lines and not lines[-1]:
        lines.pop()
    return "\n".join(lines)


def _score_sheet(violations):
    weighted = 0
    for violation in violations:
        weighted += VIOLATION_WEIGHTS.get(violation.get("code"), 5)
    return max(0, 100 - weighted)


def _load_from_xlsx(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheets = [_normalize_sheet_name(name) for name in workbook.sheetnames if _looks_like_weekly_plan(name)]
    data = []
    for name in sheets:
        sheet = workbook[name]
        workouts = _parse_workouts_from_xlsx_sheet(sheet)
        data.append((name, workouts))
    return data


def _load_from_google(config):
    from src.sheets_reader import SheetsReader

    reader = SheetsReader(
        credentials_file=config["google_sheets"]["credentials_file"],
        spreadsheet_id=config["google_sheets"]["spreadsheet_id"],
        sheet_name=config["google_sheets"]["sheet_name"],
        service_account_file=config.get("google_sheets", {}).get("service_account_file"),
    )
    reader.authenticate()
    sheets = list(reversed(reader.get_all_weekly_plan_sheets()))
    data = []
    for sheet_name in sheets:
        reader.sheet_name = sheet_name
        workouts = reader.read_workout_history(num_recent_workouts=20)
        data.append((sheet_name, workouts))
    return data


def run_backtest(sheet_data, limit=None):
    results = []
    code_counter = Counter()

    selected = sheet_data[:limit] if limit and limit > 0 else sheet_data
    for sheet_name, workouts in selected:
        markdown = _workouts_to_markdown(workouts)
        validation = validate_plan(markdown, progression_directives=[])
        violations = validation["violations"]
        for violation in violations:
            code_counter[violation["code"]] += 1

        results.append(
            {
                "sheet_name": sheet_name,
                "summary": validation["summary"],
                "violation_count": len(violations),
                "score": _score_sheet(violations),
                "violations": violations,
            }
        )

    aggregate_score = round(sum(result["score"] for result in results) / len(results), 2) if results else 0.0
    return {
        "generated_at": datetime.now().isoformat(),
        "sheet_count": len(results),
        "aggregate_score": aggregate_score,
        "violation_code_counts": dict(code_counter),
        "results": results,
    }


def write_reports(report, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = os.path.join(output_dir, f"generation_backtest_{stamp}.json")
    md_path = os.path.join(output_dir, f"generation_backtest_{stamp}.md")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)

    lines = [
        "# Generation Quality Backtest",
        "",
        f"- Generated at: {report['generated_at']}",
        f"- Sheets tested: {report['sheet_count']}",
        f"- Aggregate score: {report['aggregate_score']}",
        "",
        "## Violation Totals by Code",
    ]
    if report["violation_code_counts"]:
        for code, count in sorted(report["violation_code_counts"].items(), key=lambda item: (-item[1], item[0])):
            lines.append(f"- `{code}`: {count}")
    else:
        lines.append("- No violations detected.")

    lines.append("")
    lines.append("## Per-Sheet Results")
    for result in report["results"]:
        lines.append(f"- `{result['sheet_name']}` | score {result['score']} | violations {result['violation_count']}")
        for violation in result["violations"][:8]:
            day = violation.get("day") or "N/A"
            exercise = violation.get("exercise") or "N/A"
            lines.append(f"  - `{violation['code']}` | {day} | {exercise} | {violation['message']}")
        if result["violation_count"] > 8:
            lines.append(f"  - ... {result['violation_count'] - 8} more")

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    return json_path, md_path


def parse_args():
    parser = argparse.ArgumentParser(description="Backtest weekly plan rule compliance.")
    parser.add_argument(
        "--xlsx-path",
        type=str,
        default="",
        help="Optional local XLSX export path. If omitted, Google Sheets API is used.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional number of most-recent sheets to evaluate.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="output/backtest",
        help="Directory for markdown/json reports.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    if args.xlsx_path:
        sheet_data = _load_from_xlsx(args.xlsx_path)
    else:
        with open("config.yaml", "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        sheet_data = _load_from_google(config)

    if not sheet_data:
        raise RuntimeError("No weekly plan sheets found for backtest.")

    report = run_backtest(sheet_data, limit=args.limit)
    json_path, md_path = write_reports(report, args.output_dir)

    print(f"Backtest complete. Sheets: {report['sheet_count']} | Aggregate score: {report['aggregate_score']}")
    print(f"JSON report: {json_path}")
    print(f"Markdown report: {md_path}")


if __name__ == "__main__":
    main()
